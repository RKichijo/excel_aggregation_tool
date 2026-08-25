import pandas as pd
import glob
import unicodedata
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

output_dir = "output"
os.makedirs(output_dir, exist_ok=True)

today = datetime.now().strftime("%y%m%d")
output_file = os.path.join(output_dir, f"集計結果_{today}.xlsx")

def load_excel_files():
    files = glob.glob("input/*.xlsx")

    if len(files) == 0:
        raise FileNotFoundError("Excelファイルが見つかりません")

    success_count = 0
    skip_count = 0
    data_list = []

    required_columns = ["店舗", "担当者", "商品名", "数量", "売上金額"]

    for file in files:
        try:
            df = pd.read_excel(file)
        except Exception as e:
            print(f"{file}：読み込みに失敗しました")
            print(f"エラー内容：{e}")
            skip_count += 1
            continue

        if not all(column in df.columns for column in required_columns):
            print(f"{file}：必要な列がありません")
            skip_count += 1
            continue

        data_list.append(df)
        success_count += 1

    if len(data_list) == 0:
        raise ValueError("正常に読み込めるExcelファイルがありません")

    all_data = pd.concat(data_list, ignore_index=True)

    return all_data, files, success_count, skip_count

def aggregate_data(all_data):
    store_sales = (
        all_data.groupby("店舗")["売上金額"]
        .sum()
        .sort_values(ascending=False)
    )

    staff_sales = (
        all_data.groupby(["担当者", "商品名"])[["数量", "売上金額"]]
        .sum()
        .sort_values("売上金額", ascending=False)
    )

    return store_sales, staff_sales


def export_excel(all_data, store_sales, staff_sales):

    with pd.ExcelWriter(output_file) as writer:
        all_data.to_excel(writer, sheet_name="全データ", index=False)
        store_sales.to_excel(writer, sheet_name="店舗ごと売上金額")
        staff_sales.to_excel(writer, sheet_name="担当者ごと商品別集計")

def get_display_width(value):
    width = 0

    for char in str(value):
        if unicodedata.east_asian_width(char) in ["F", "W", "A"]:
            width += 2
        else:
            width += 1

    return width


def format_excel():
    wb = load_workbook(output_file)

    for ws in wb.worksheets:

        # 見出し装飾
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                fill_type="darkTrellis",
                fgColor="D9EAF7"
            )

        # 列幅調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    cell_length = get_display_width(cell.value)
                    max_length = max(max_length, cell_length)
            ws.column_dimensions[column_letter].width = max_length + 2

        # 売上金額をカンマ区切り
        for cell in ws[1]:
            if cell.value == "売上金額":
                column_number = int(cell.column)

                for row in range(2, ws.max_row + 1):
                    ws.cell(
                        row=row,
                        column=column_number
                    ).number_format = '#,##0'

    wb.save(output_file)

# ① Excelを読み込む
all_data, files, success_count, skip_count = load_excel_files()

# ② 読み込んだデータを集計する
store_sales, staff_sales = aggregate_data(all_data)

# ③ 集計データをエクスポートする
export_excel(all_data, store_sales, staff_sales)

# ④ エクスポートデータを成形する
format_excel()

print(f"対象ファイル：{len(files)}件")
print(f"正常読込：{success_count}件")
print(f"スキップ：{skip_count}件")
print("実行完了")